"""
Services Premium pour les fonctionnalités avancées
"""
from datetime import datetime
from flask import url_for
from app.models import Payment, Tenant
import io


def export_payments_to_excel(user):
    """
    Exporte tous les paiements de l'utilisateur vers un fichier Excel.
    Fonctionnalité réservée au plan Premium.
    
    Args:
        user: Instance de User
        
    Returns:
        BytesIO: Fichier Excel en mémoire
    """
    try:
        # Import conditionnel de openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl n'est pas installé. Installez-le avec: pip install openpyxl")
    
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique Paiements"
    
    # En-têtes
    headers = ['Date', 'Période', 'Locataire', 'Téléphone', 'Immeuble', 
               'Appartement', 'Montant (FCFA)', 'WhatsApp Envoyé', 'Rappel Envoyé']
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Récupérer tous les paiements de l'utilisateur
    all_payments = []
    for property in user.properties:
        for unit in property.units:
            for tenant in unit.tenants:
                for payment in tenant.payments:
                    all_payments.append({
                        'payment': payment,
                        'tenant': tenant,
                        'unit': unit,
                        'property': property
                    })
    
    # Trier par date décroissante
    all_payments.sort(key=lambda x: x['payment'].date_paid, reverse=True)
    
    # Remplir les données
    for row_num, item in enumerate(all_payments, 2):
        payment = item['payment']
        tenant = item['tenant']
        unit = item['unit']
        property = item['property']
        
        ws.cell(row=row_num, column=1).value = payment.date_paid.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=2).value = payment.period
        ws.cell(row=row_num, column=3).value = tenant.full_name
        ws.cell(row=row_num, column=4).value = tenant.phone
        ws.cell(row=row_num, column=5).value = property.name
        ws.cell(row=row_num, column=6).value = unit.door_number
        ws.cell(row=row_num, column=7).value = payment.amount
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        ws.cell(row=row_num, column=8).value = 'Oui' if payment.whatsapp_sent else 'Non'
        ws.cell(row=row_num, column=9).value = 'Oui' if payment.reminder_sent else 'Non'
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(str(headers[col-1]))
        
        for row in range(2, len(all_payments) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Ajouter une ligne de totaux
    if all_payments:
        total_row = len(all_payments) + 2
        ws.cell(row=total_row, column=6).value = "TOTAL:"
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).value = f"=SUM(G2:G{total_row-1})"
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).number_format = '#,##0'
    
    # Sauvegarder dans un buffer mémoire
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def send_whatsapp_receipt(payment, tenant):
    """
    Génère un lien WhatsApp pour envoyer la quittance au locataire.
    Fonctionnalité Premium.
    
    Args:
        payment: Instance de Payment
        tenant: Instance de Tenant
        
    Returns:
        str: URL WhatsApp pré-remplie
    """
    from urllib.parse import quote
    
    # Générer le lien vers le PDF
    pdf_url = url_for('finances.download_receipt', payment_id=payment.id, _external=True)
    
    # Formater le montant
    amount_fmt = "{:,.0f}".format(payment.amount).replace(',', ' ')
    
    # Message personnalisé
    msg_text = (f"Bonjour {tenant.full_name}, "
                f"votre paiement de {amount_fmt} FCFA pour la période {payment.period} "
                f"a bien été reçu. Merci ! 🏠\n\n"
                f"Téléchargez votre quittance ici : {pdf_url}")
    
    # Nettoyer le numéro de téléphone
    phone = tenant.phone.replace(' ', '').replace('+', '')
    
    # Générer le lien WhatsApp
    whatsapp_url = f"https://wa.me/{phone}?text={quote(msg_text)}"
    
    return whatsapp_url


def send_whatsapp_reminder(tenant, amount, period):
    """
    Génère un lien WhatsApp pour envoyer un rappel de paiement.
    Fonctionnalité Premium.
    
    Args:
        tenant: Instance de Tenant
        amount: Montant du loyer
        period: Période concernée (ex: "2023-12")
        
    Returns:
        str: URL WhatsApp pré-remplie
    """
    from urllib.parse import quote
    
    # Formater le montant
    amount_fmt = "{:,.0f}".format(amount).replace(',', ' ')
    
    # Message de rappel courtois mais ferme
    msg_text = (f"Bonjour {tenant.full_name}, "
                f"sauf erreur de notre part, nous n'avons pas encore reçu votre loyer de {amount_fmt} FCFA "
                f"pour la période {period}. "
                f"Merci de régulariser votre situation dès que possible. 🏠")
    
    # Nettoyer le numéro de téléphone
    phone = tenant.phone.replace(' ', '').replace('+', '')
    
    # Générer le lien WhatsApp
    whatsapp_url = f"https://wa.me/{phone}?text={quote(msg_text)}"
    
    return whatsapp_url


def get_late_tenants(user):
    """
    Identifie les locataires qui n'ont pas payé pour le mois en cours.
    Fonctionnalité Premium.
    
    Args:
        user: Instance de User
        
    Returns:
        list: Liste de dictionnaires avec tenant, unit, property, amount_due
    """
    late_tenants = []
    # Période actuelle (Mois en cours)
    current_period = datetime.now().strftime('%Y-%m')
    
    # Si on est avant le 5 du mois, on vérifie peut-être le mois précédent ?
    # Pour simplifier : on vérifie toujours le mois en cours par défaut
    
    for prop in user.properties:
        for unit in prop.units:
            # On ne vérifie que les appartements occupés
            tenant = unit.current_tenant
            if tenant:
                # Vérifier s'il existe un paiement pour ce mois
                # Note: On utilise une requête directe pour être plus efficace
                has_paid = False
                for payment in tenant.payments:
                    if payment.period == current_period:
                        has_paid = True
                        break
                
                if not has_paid:
                    late_tenants.append({
                        'tenant': tenant,
                        'unit': unit,
                        'property': prop,
                        'amount_due': unit.rent_amount,
                        'period': current_period
                    })
    
    return late_tenants


def get_payment_statistics(user):
    """
    Calcule des statistiques avancées sur les paiements.
    Fonctionnalité Premium - pour le dashboard analytique.
    
    Args:
        user: Instance de User
        
    Returns:
        dict: Statistiques diverses
    """
    from collections import defaultdict
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    
    # Initialiser les 12 derniers mois à 0
    monthly_revenue = {}
    today = datetime.now()
    for i in range(11, -1, -1):
        date = today - relativedelta(months=i)
        key = date.strftime('%Y-%m')
        monthly_revenue[key] = 0.0
    
    total_payments = 0
    total_revenue = 0
    
    for property in user.properties:
        for unit in property.units:
            for tenant in unit.tenants:
                for payment in tenant.payments:
                    # Grouper par mois
                    month_key = payment.period  # Format "2023-11"
                    # On ne garde que si c'est dans notre fenêtre de 12 mois ou futur
                    if month_key in monthly_revenue:
                        monthly_revenue[month_key] += payment.amount
                    # Ou on l'ajoute si on veut tout l'historique (mais le chart risque d'être long)
                    # Ici on privilégie la vue "12 derniers mois" pour le dashboard
                    
                    total_payments += 1
                    total_revenue += payment.amount
    
    # Trier par mois (déjà fait par la boucle d'init, mais on s'assure)
    sorted_months = sorted(monthly_revenue.items())
    
    # Calculer le taux de recouvrement
    total_units = user.get_total_units()
    occupied_units = sum(1 for prop in user.properties 
                        for unit in prop.units if unit.current_tenant)
    
    # Paiements moyens
    avg_payment = total_revenue / total_payments if total_payments > 0 else 0
    
    return {
        'monthly_revenue': dict(sorted_months),
        'total_payments': total_payments,
        'total_revenue': total_revenue,
        'avg_payment': avg_payment,
        'collection_rate': (occupied_units / total_units * 100) if total_units > 0 else 0
    }
